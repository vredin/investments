"""Unit tests for Freedom Finance Excel parsers.

Tests parse_freedom_portfolio_xlsx and parse_freedom_trades_xlsx without a DB.
Security edge cases: zip-bomb guard, magic bytes, missing columns, empty file.
"""

import io
import zipfile
from datetime import date

import openpyxl
import pytest

from app.services.ingestion.broker import BrokerSyncError
from app.services.ingestion.freedom import (
    parse_freedom_portfolio_xlsx,
    parse_freedom_trades_xlsx,
)

_PORTFOLIO_FIXTURE = "tests/fixtures/freedom_portfolio_sample.xlsx"
_TRADES_FIXTURE = "tests/fixtures/freedom_trades_sample.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_portfolio_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Тікер", "К-ть", "Ціна входу", "Ціна", "Вартість", "Частка (%)", "Прибуток", "Приріст"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_trades_xlsx(rows: list[list]) -> bytes:

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Номер", "Дата", "Розрахунки", "Тікер", "Операція", "Quantity", "Ціна", "Сума", "Прибуток", "Плата"])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# parse_freedom_portfolio_xlsx — happy path
# ---------------------------------------------------------------------------


def test_portfolio_parses_fixture_file():
    with open(_PORTFOLIO_FIXTURE, "rb") as f:
        rows = parse_freedom_portfolio_xlsx(f.read())
    assert len(rows) == 1  # cash row USD/EUR must be skipped
    assert rows[0].ticker == "AAPL.US"


def test_portfolio_correct_field_mapping():
    content = _make_portfolio_xlsx([["ZIM.US", 7, 13.84, 26.05, 182.35, 93.11, 85.5, 88.28]])
    rows = parse_freedom_portfolio_xlsx(content)
    assert len(rows) == 1
    r = rows[0]
    assert r.ticker == "ZIM.US"
    assert r.qty == 7.0
    assert r.market_value == 182.35
    assert abs(r.cost_basis - 7 * 13.84) < 0.01
    assert r.broker == "freedom"
    assert r.currency == "USD"


def test_portfolio_cash_rows_skipped():
    content = _make_portfolio_xlsx([
        ["USD/EUR", 100, 0.9, 0.91, 91.0, 25.0, 1.0, 1.1],
        ["EUR/USD", 50, 1.1, 1.09, 54.5, 10.0, 0.0, 0.0],
    ])
    rows = parse_freedom_portfolio_xlsx(content)
    assert rows == []


def test_portfolio_returns_empty_for_header_only():
    content = _make_portfolio_xlsx([])
    rows = parse_freedom_portfolio_xlsx(content)
    assert rows == []


# ---------------------------------------------------------------------------
# parse_freedom_portfolio_xlsx — error cases
# ---------------------------------------------------------------------------


def test_portfolio_missing_required_column_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["WrongCol1", "WrongCol2"])
    ws.append(["ZIM.US", 7])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(BrokerSyncError, match="missing columns"):
        parse_freedom_portfolio_xlsx(buf.getvalue())


def test_portfolio_rejects_non_xlsx_magic_bytes():
    with pytest.raises(BrokerSyncError, match="magic bytes"):
        parse_freedom_portfolio_xlsx(b"This is not an xlsx file at all")


def test_portfolio_rejects_file_over_5mb():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as z:
        z.writestr("padding.bin", b"X" * (5 * 1024 * 1024 + 100))
    with pytest.raises(BrokerSyncError, match="too large"):
        parse_freedom_portfolio_xlsx(buf.getvalue())


# ---------------------------------------------------------------------------
# parse_freedom_trades_xlsx — happy path
# ---------------------------------------------------------------------------


def test_trades_parses_fixture_file():
    with open(_TRADES_FIXTURE, "rb") as f:
        rows = parse_freedom_trades_xlsx(f.read())
    assert len(rows) == 1  # USD/EUR row skipped
    assert rows[0].ticker == "AAPL.US"


def test_trades_correct_field_mapping():
    from datetime import datetime

    content = _make_trades_xlsx([
        [527716496, datetime(2025, 4, 23, 13, 30), datetime(2025, 4, 24), "ZIM.US", " Купівля ", 7, 13.84, 96.85, 0.0, 1.85]
    ])
    rows = parse_freedom_trades_xlsx(content)
    assert len(rows) == 1
    r = rows[0]
    assert r.ticker == "ZIM.US"
    assert r.txn_type == "BUY"
    assert r.qty == 7.0
    assert r.price == 13.84
    assert r.commission == 1.85
    assert r.ibkr_txn_id == "freedom_527716496"
    assert r.trade_date == date(2025, 4, 23)
    assert r.broker == "freedom"


def test_trades_sell_operation_mapped():
    from datetime import datetime

    content = _make_trades_xlsx([
        [111, datetime(2025, 6, 1, 9, 0), datetime(2025, 6, 2), "ZIM.US", " Продаж ", 3, 25.0, 75.0, 10.0, 0.5]
    ])
    rows = parse_freedom_trades_xlsx(content)
    assert rows[0].txn_type == "SELL"


def test_trades_cash_rows_skipped():
    from datetime import datetime

    content = _make_trades_xlsx([
        [999, datetime(2025, 9, 11), datetime(2025, 9, 12), "USD/EUR", " Продаж ", 7.03, 0.85, 6.0, 0.0, 0.0]
    ])
    rows = parse_freedom_trades_xlsx(content)
    assert rows == []


def test_trades_missing_required_column_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["BadCol1", "BadCol2"])
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(BrokerSyncError, match="missing columns"):
        parse_freedom_trades_xlsx(buf.getvalue())


def test_trades_rejects_non_xlsx_magic_bytes():
    with pytest.raises(BrokerSyncError, match="magic bytes"):
        parse_freedom_trades_xlsx(b"<xml>not xlsx</xml>")
