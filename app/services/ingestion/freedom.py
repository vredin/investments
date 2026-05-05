"""Freedom Finance adapter.

REST API (tradernet.com) does not expose portfolio/positions via HTTP —
those are WebSocket-only (notifyPortfolio). Import is done via Excel exports
from the Freedom24 web UI (Профіль → Портфель → Експорт в Excel).

parse_freedom_portfolio_xlsx  — парсить "Відкриті позиції" Excel
parse_freedom_trades_xlsx     — парсить "Угоди" Excel
"""

import hashlib
import hmac
import io
import json
import logging
import time
from datetime import date, timedelta

import openpyxl
import requests

from app.config import get_settings
from app.services.ingestion.broker import (
    BrokerSyncError,
    PositionRow,
    TransactionRow,
    upsert_positions,
    upsert_transactions,
)

logger = logging.getLogger(__name__)

_API_URL = "https://tradernet.com/api"
_TIMEOUT = 30
_MAX_RETRIES = 2
_BROKER = "freedom"

# TraderNet portfolio API method name (REST endpoint)
_METHOD_PORTFOLIO = "getPortfolio"
# TraderNet trade history API method name
_METHOD_TRANSACTIONS = "getTransactionHistory"


def _send_request(method: str, params: dict | None = None) -> dict:
    """Send HMAC-signed request to TraderNet REST API.

    Raises BrokerSyncError on auth failure, timeout, or unexpected HTTP response.
    Never logs the response payload (PII/financial data rule).
    """
    settings = get_settings()
    if not settings.FREEDOM_PUBLIC_KEY or not settings.FREEDOM_PRIVATE_KEY:
        raise BrokerSyncError("Freedom Finance API keys not configured in environment")

    payload = json.dumps(
        {"method": method, "apiKey": settings.FREEDOM_PUBLIC_KEY, **(params or {})},
        separators=(",", ":"),
        sort_keys=True,
    )
    sig = hmac.new(
        settings.FREEDOM_PRIVATE_KEY.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()

    last_exc: Exception | None = None
    for attempt in range(_MAX_RETRIES + 1):
        try:
            resp = requests.post(
                _API_URL,
                data={"data": payload, "sig": sig},
                timeout=_TIMEOUT,
            )
            resp.raise_for_status()
            return resp.json()
        except requests.Timeout as exc:
            last_exc = exc
            logger.warning("Freedom API timeout (attempt %d/%d)", attempt + 1, _MAX_RETRIES + 1)
        except requests.HTTPError as exc:
            raise BrokerSyncError(
                f"Freedom API HTTP {exc.response.status_code}: {exc.response.reason}"
            ) from exc
        except requests.RequestException as exc:
            raise BrokerSyncError(f"Freedom API request failed: {exc}") from exc

        if attempt < _MAX_RETRIES:
            time.sleep(2**attempt)  # 1s, 2s

    raise BrokerSyncError(
        f"Freedom API timed out after {_MAX_RETRIES + 1} attempts ({_TIMEOUT}s each)"
    ) from last_exc


def _normalize_freedom_position(item: dict) -> PositionRow | None:
    """Normalize a TraderNet portfolio position dict to PositionRow. Returns None on bad data."""
    ticker = item.get("i", "")
    if not ticker:
        logger.warning("Freedom position item missing ticker field, skipping")
        return None
    try:
        return PositionRow(
            ticker=str(ticker),
            qty=float(item.get("q") or 0),
            market_value=float(item.get("s") or 0),
            cost_basis=float(item.get("open_bal") or 0),
            currency=str(item.get("curr") or "USD"),
            broker=_BROKER,
        )
    except (ValueError, TypeError) as exc:
        logger.warning("Skipping Freedom position %s: %s", ticker, exc)
        return None


def _normalize_freedom_transaction(item: dict) -> TransactionRow | None:
    """Normalize a TraderNet order dict to TransactionRow. Returns None on bad data."""
    order_id = str(item.get("id") or "")
    ticker = str(item.get("instr") or "")
    if not ticker:
        logger.warning("Freedom transaction missing ticker, skipping")
        return None

    try:
        ts = item.get("date") or 0
        trade_date = date.fromtimestamp(float(ts)) if ts else date.today()

        oper = int(item.get("oper") or 1)
        txn_type = "BUY" if oper in (1, 2) else "SELL"

        qty = abs(float(item.get("q") or 0))
        price = float(item.get("p") or 0)

        ibkr_txn_id = (
            f"freedom_{order_id}"
            if order_id
            else f"freedom_{trade_date.isoformat()}_{ticker}_{qty}"
        )

        return TransactionRow(
            ibkr_txn_id=ibkr_txn_id,
            ticker=ticker,
            trade_date=trade_date,
            txn_type=txn_type,
            qty=qty,
            price=price,
            amount=abs(qty * price),
            commission=float(item.get("commission") or 0),
            currency=str(item.get("curr") or "USD"),
            broker=_BROKER,
        )
    except (ValueError, TypeError, OSError) as exc:
        logger.warning("Skipping Freedom transaction %s: %s", order_id, exc)
        return None


_MAX_XLSX_BYTES = 5 * 1024 * 1024  # 5 MB — guards against zip-bomb DoS
_XLSX_MAGIC = b"PK\x03\x04"        # xlsx is a zip; all valid files start with this


def _validate_xlsx_bytes(content: bytes, label: str) -> None:
    if len(content) > _MAX_XLSX_BYTES:
        raise BrokerSyncError(f"{label}: file too large (max 5 MB)")
    if not content.startswith(_XLSX_MAGIC):
        raise BrokerSyncError(f"{label}: not a valid .xlsx file (wrong magic bytes)")


def parse_freedom_portfolio_xlsx(file_content: bytes) -> list[PositionRow]:
    """Parse Freedom24 'Відкриті позиції' Excel export into PositionRow list.

    Expected columns (row 1 header):
      Тікер | К-ть | Ціна входу | Ціна | Вартість | Частка (%) | Прибуток | Приріст
    """
    _validate_xlsx_bytes(file_content, "Portfolio Excel")
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_content), data_only=True, keep_links=False
        )
    except Exception as exc:
        raise BrokerSyncError(f"Cannot open portfolio Excel: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BrokerSyncError("Portfolio Excel is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    required = {"Тікер", "К-ть", "Ціна входу", "Вартість"}
    missing = required - set(header)
    if missing:
        raise BrokerSyncError(f"Portfolio Excel missing columns: {missing}")

    idx = {name: i for i, name in enumerate(header)}
    result: list[PositionRow] = []

    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        ticker = str(row[idx["Тікер"]] or "").strip()
        if not ticker or "/" in ticker:
            # Skip cash rows like USD/EUR
            continue
        try:
            qty = float(row[idx["К-ть"]] or 0)
            entry_price = float(row[idx["Ціна входу"]] or 0)
            market_value = float(row[idx["Вартість"]] or 0)
            cost_basis = round(qty * entry_price, 6)
            # Determine currency from ticker suffix (.US → USD, else USD default)
            currency = "USD" if ticker.endswith(".US") else "USD"
            result.append(
                PositionRow(
                    ticker=ticker,
                    qty=qty,
                    market_value=market_value,
                    cost_basis=cost_basis,
                    currency=currency,
                    broker=_BROKER,
                )
            )
        except (ValueError, TypeError) as exc:
            logger.warning("Skipping portfolio row %s: %s", ticker, exc)

    return result


def parse_freedom_trades_xlsx(file_content: bytes) -> list[TransactionRow]:
    """Parse Freedom24 'Угоди' Excel export into TransactionRow list.

    Expected columns (row 1 header):
      Номер | Дата | Розрахунки | Тікер | Операція | Quantity | Ціна | Сума | Прибуток | Плата
    """
    _validate_xlsx_bytes(file_content, "Trades Excel")
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_content), data_only=True, keep_links=False
        )
    except Exception as exc:
        raise BrokerSyncError(f"Cannot open trades Excel: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BrokerSyncError("Trades Excel is empty")

    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    required = {"Номер", "Дата", "Тікер", "Операція", "Quantity", "Ціна", "Сума"}
    missing = required - set(header)
    if missing:
        raise BrokerSyncError(f"Trades Excel missing columns: {missing}")

    idx = {name: i for i, name in enumerate(header)}
    result: list[TransactionRow] = []

    for row in rows[1:]:
        if not any(v is not None for v in row):
            continue
        ticker = str(row[idx["Тікер"]] or "").strip()
        if not ticker or "/" in ticker:
            # Skip cash conversion rows like USD/EUR
            continue
        try:
            order_id = str(row[idx["Номер"]] or "").strip()
            raw_date = row[idx["Дата"]]
            if hasattr(raw_date, "date"):
                trade_date = raw_date.date()
            else:
                from datetime import datetime
                trade_date = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()

            operation = str(row[idx["Операція"]] or "").strip().lower()
            txn_type = "BUY" if "купівля" in operation or "buy" in operation else "SELL"

            qty = abs(float(row[idx["Quantity"]] or 0))
            price = float(row[idx["Ціна"]] or 0)
            amount = abs(float(row[idx["Сума"]] or 0))
            commission = float(row[idx.get("Плата", -1)] or 0) if "Плата" in idx else 0.0
            currency = "USD" if ticker.endswith(".US") else "USD"

            ibkr_txn_id = f"freedom_{order_id}" if order_id else (
                f"freedom_{trade_date.isoformat()}_{ticker}_{qty}"
            )

            result.append(
                TransactionRow(
                    ibkr_txn_id=ibkr_txn_id,
                    ticker=ticker,
                    trade_date=trade_date,
                    txn_type=txn_type,
                    qty=qty,
                    price=price,
                    amount=amount,
                    commission=commission,
                    currency=currency,
                    broker=_BROKER,
                )
            )
        except (ValueError, TypeError, AttributeError) as exc:
            logger.warning("Skipping trade row %s: %s", row, exc)

    return result


def import_freedom_portfolio_xlsx(file_content: bytes, db) -> int:
    """Parse portfolio Excel and upsert positions. Returns count upserted."""
    rows = parse_freedom_portfolio_xlsx(file_content)
    count = upsert_positions(db, rows)
    logger.info("Imported %d Freedom Finance positions from Excel", count)
    return count


def import_freedom_trades_xlsx(file_content: bytes, db) -> int:
    """Parse trades Excel and upsert transactions. Returns count upserted."""
    rows = parse_freedom_trades_xlsx(file_content)
    count = upsert_transactions(db, rows)
    logger.info("Imported %d Freedom Finance transactions from Excel", count)
    return count


def sync_freedom_positions(db) -> int:
    """Fetch Freedom Finance portfolio and upsert positions. Returns count of rows upserted.

    Raises BrokerSyncError on API failure. Never calls db.commit() — caller owns transaction.
    """
    data = _send_request(_METHOD_PORTFOLIO)

    # Expected response shape: {"result": {"pos": [...]}} or {"pos": [...]}
    positions_raw = (
        data.get("result", {}).get("pos")
        or data.get("pos")
        or []
    )

    rows = [r for item in positions_raw if (r := _normalize_freedom_position(item)) is not None]
    count = upsert_positions(db, rows)
    logger.info("Synced %d positions from Freedom Finance", count)
    return count


def sync_freedom_transactions(db) -> int:
    """Fetch Freedom Finance trade history (last 365 days) and upsert. Returns count upserted.

    Raises BrokerSyncError on API failure. Never calls db.commit() — caller owns transaction.
    """
    today = date.today()
    from_date = (today - timedelta(days=365)).isoformat()
    to_date = today.isoformat()

    data = _send_request(_METHOD_TRANSACTIONS, {"from": from_date, "to": to_date})

    transactions_raw = (
        data.get("result", {}).get("orders")
        or data.get("orders")
        or data.get("result", {}).get("transactions")
        or data.get("transactions")
        or []
    )

    rows = [
        r
        for item in transactions_raw
        if (r := _normalize_freedom_transaction(item)) is not None
    ]
    count = upsert_transactions(db, rows)
    logger.info("Synced %d transactions from Freedom Finance", count)
    return count
